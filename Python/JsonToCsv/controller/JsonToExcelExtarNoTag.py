import json
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

# Folder paths for JSON and Excel
json_folder_path = "../data"  # Adjust path for JSON files
excel_output_folder = "../csv"  # Adjust path for Excel output

# Create the folder for Excel output if it doesn't exist
os.makedirs(excel_output_folder, exist_ok=True)

# Define border style for cells
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# Create a new Excel workbook
wb = Workbook()

# Iterate through files in the JSON folder
for idx, filename in enumerate(os.listdir(json_folder_path), 1):
    # Process only .json files
    if filename.endswith(".json"):
        file_path = os.path.join(json_folder_path, filename)

        # Check if the file is empty
        if os.path.getsize(file_path) == 0:
            print(f"Skipping empty file: {filename}")
            continue  # Skip empty file

        # Try to load the JSON data, with error handling
        try:
            with open(file_path, "r") as json_file:
                data = json.load(json_file)

            # Use the filename (without the .json extension) as the sheet name
            sheet_name = filename.replace(".json", "")  # Remove the .json extension
            ws = wb.create_sheet(title=sheet_name)

            # Write headers
            ws.append(["Service", "ResourceARN", "Tags"])

            # Prepare the data to be written into the Excel sheet
            all_rows = []
            for resource in data.get("ResourceTagMappingList", []):
                resource_arn = resource["ResourceARN"]
                service_name = resource_arn.split(":")[2]  # Extract service name from ARN
                tags = resource.get("Tags", [])
                
                # Collect tags or mark "No Tag" if not available
                if not tags:  # Only select rows with "No Tag"
                    tag_str = "No Tag"
                    all_rows.append([service_name, resource_arn, tag_str])

            # Write data into the Excel sheet, but only the rows with "No Tag"
            for row in all_rows:
                ws.append(row)

            # Apply borders to all cells
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=len(all_rows) + 1, min_col=1, max_col=3), start=2):
                for cell in row:
                    cell.border = border

            # Set 'wrap text' and 'vertical alignment' to top for all columns (A, B, C)
            for col in range(1, 4):  # Columns A, B, C (1, 2, 3)
                for row in ws.iter_rows(min_row=2, max_row=len(all_rows) + 1, min_col=col, max_col=col):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")  # Align text to top

            # Set column widths for A, B, and C
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 26
            ws.column_dimensions["C"].width = 40

            print(f"Data from {filename} has been written to {sheet_name}")

        except json.JSONDecodeError as e:
            print(f"Error decoding JSON from {filename}: {e}")
        except Exception as e:
            print(f"Unexpected error with file {filename}: {e}")

# Delete the default empty sheet created by openpyxl
del wb["Sheet"]

# Save the Excel file
excel_filename = "combined_data_no_tag.xlsx"
excel_file_path = os.path.join(excel_output_folder, excel_filename)
wb.save(excel_file_path)
print(f"All data has been written to {excel_filename}")
