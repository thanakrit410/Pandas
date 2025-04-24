import json
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment

# กำหนด path ของโฟลเดอร์ JSON และไฟล์ Excel
json_folder_path = "../data5"  # ปรับ path ของ JSON
excel_output_folder = "../csv"  # ปรับ path ของ Excel output

# สร้างโฟลเดอร์สำหรับ Excel output ถ้ายังไม่มี
os.makedirs(excel_output_folder, exist_ok=True)

# สไตล์เส้นขอบตาราง
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# สีแดงสำหรับข้อมูลที่ขาดหายไป
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# สีเขียวสำหรับ GeoRestrictionType ที่ไม่ใช่ 'none'
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# สีขาวสำหรับ GeoRestrictionType ที่เป็น 'none'
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# สร้าง Workbook สำหรับ Excel
wb = Workbook()

# ลบ Sheet เริ่มต้นที่ว่าง
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# วนลูปอ่านไฟล์ JSON ที่อยู่ในโฟลเดอร์
for filename in os.listdir(json_folder_path):
    if filename.endswith(".json"):
        file_path = os.path.join(json_folder_path, filename)

        # เช็คว่าไฟล์ว่างหรือไม่
        if os.path.getsize(file_path) == 0:
            print(f"Skipping empty file: {filename}")
            continue

        # โหลด JSON
        try:
            with open(file_path, "r") as json_file:
                data = json.load(json_file)

            # ตรวจสอบว่ามีข้อมูลหรือไม่
            if not data:
                print(f"No data in file: {filename}")
                continue

            # ใช้ชื่อไฟล์ (ตัด .json ออก) เป็นชื่อ Sheet
            sheet_name = filename.replace(".json", "")
            ws = wb.create_sheet(title=sheet_name)

            # หาหัวข้อคอลัมน์จาก JSON อัตโนมัติ
            columns = list(data[0].keys()) if data else []

            # เขียน Header
            ws.append(columns)

            # ใส่ข้อมูลจาก JSON
            for item in data:
                # แก้ไขค่า 'None' หรือ ['null'] ที่ไม่สามารถแสดงใน Excel ได้
                if 'Countries' in item:
                    # ถ้าเป็น [] หรือ ['null'] จะเปลี่ยนเป็น 'N/A'
                    if item['Countries'] == [] or item['Countries'] == ['null']:
                        item['Countries'] = "N/A"
                    else:
                        item['Countries'] = ", ".join([x for x in item['Countries'] if x != 'null'])

                row_data = [item.get(col, "N/A") for col in columns]
                ws.append(row_data)

            # จัดรูปแบบตาราง
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    # ไฮไลท์สีแดงถ้าเป็น "N/A" หรือไม่มีข้อมูล
                    if cell.value in ["N/A", None, ""]:
                        cell.fill = red_fill

                    # ถ้า GeoRestrictionType เป็น 'none' ให้ใช้สีขาว
                    if cell.column == columns.index("GeoRestrictionType") + 1:
                        if cell.value == "none":
                            cell.fill = white_fill
                        else:
                            cell.fill = green_fill

            # ปรับขนาดคอลัมน์ให้อัตโนมัติ
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                ws.column_dimensions[col[0].column_letter].width = max_length

            print(f"Processed: {filename} → {sheet_name}")

        except json.JSONDecodeError as e:
            print(f"Error decoding JSON from {filename}: {e}")
        except Exception as e:
            print(f"Unexpected error with file {filename}: {e}")

# บันทึกไฟล์ Excel
excel_filename = "cloudfront_instances.xlsx"
excel_file_path = os.path.join(excel_output_folder, excel_filename)
wb.save(excel_file_path)
print(f"Excel file saved: {excel_filename}")
