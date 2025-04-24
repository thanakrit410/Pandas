import json
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment

# กำหนด path ของโฟลเดอร์ JSON และไฟล์ Excel
json_folder_path = "../data3"  # ปรับ path ของ JSON
excel_output_folder = "../csv"  # ปรับ path ของ Excel output

# สร้างโฟลเดอร์สำหรับ Excel output ถ้ายังไม่มี
os.makedirs(excel_output_folder, exist_ok=True)

# สไตล์เส้นขอบตาราง
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# สีแดงสำหรับข้อมูลที่ขาดหายไป
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# สร้าง Workbook สำหรับ Excel
wb = Workbook()

# ลบ Sheet เริ่มต้นที่ว่าง
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# วนลูปอ่านไฟล์ JSON ที่อยู่ในโฟลเดอร์
for filename in os.listdir(json_folder_path):
    if filename.endswith(".json"):
        file_path = os.path.join(json_folder_path, filename)

        # เช็คว่าไฟล์ว่างหรือไม่
        if os.path.getsize(file_path) == 0:
            print(f"Skipping empty file: {filename}")
            continue

        # โหลด JSON
        try:
            with open(file_path, "r") as json_file:
                data = json.load(json_file)

            # ถ้าเป็น list ของ RDS instances
            rds_instances = data if isinstance(data, list) else []

            # ข้ามไฟล์ที่ไม่มีข้อมูล RDS
            if not rds_instances:
                print(f"No RDS instance data in: {filename}")
                continue

            # ใช้ชื่อไฟล์ (ตัด .json ออก) เป็นชื่อ Sheet
            sheet_name = filename.replace(".json", "")
            ws = wb.create_sheet(title=sheet_name)

            # กำหนดคอลัมน์
            columns = ["InstanceID", "State", "InstanceType", "PublicIP", "Name", "Project"]

            # เขียน Header
            ws.append(columns)

            # ใส่ข้อมูลจาก RDS instances
            for instance in rds_instances:
                row_data = [
                    instance.get('InstanceID', 'N/A'),
                    instance.get('State', 'N/A'),
                    instance.get('InstanceType', 'N/A'),
                    instance.get('PublicIP', 'N/A'),
                    instance.get('Name', 'N/A') if instance.get('Name') else 'N/A',
                    instance.get('Project', 'N/A')
                ]
                ws.append(row_data)

            # จัดรูปแบบตาราง
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    # ไฮไลท์สีแดงถ้าเป็น "N/A" หรือไม่มีข้อมูล
                    if cell.value in ["N/A", None, ""]:
                        cell.fill = red_fill

            # ปรับขนาดคอลัมน์ให้อัตโนมัติ
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                ws.column_dimensions[col[0].column_letter].width = max_length

            print(f"Processed: {filename} → {sheet_name}")

        except json.JSONDecodeError as e:
            print(f"Error decoding JSON from {filename}: {e}")
        except Exception as e:
            print(f"Unexpected error with file {filename}: {e}")

# ถ้าไม่มี sheet ที่ถูกเพิ่มมา, ให้แสดงข้อความ
if not wb.sheetnames:
    print("No data was added to the workbook.")

# บันทึกไฟล์ Excel
excel_filename = "rds_instances.xlsx"
excel_file_path = os.path.join(excel_output_folder, excel_filename)
wb.save(excel_file_path)
print(f"Excel file saved: {excel_filename}")
