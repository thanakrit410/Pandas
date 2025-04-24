import json
import os
import boto3
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment

# 🔹 ตั้งค่า AWS S3
S3_BUCKET_NAME = "uat-autolist-resource-bucket"
S3_FILE_NAME = "ec2_instances.xlsx"

# 🔹 ตั้งค่าพาธสำหรับบันทึกไฟล์ Excel ชั่วคราวใน Lambda
TMP_FILE_PATH = "/tmp/" + S3_FILE_NAME

# 🔹 สร้าง AWS Clients
s3_client = boto3.client("s3")
ec2_client = boto3.client("ec2")

# 🔹 สไตล์ของเส้นขอบ
border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# 🔹 สีแดงสำหรับข้อมูลที่ขาด
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def get_ec2_instances():
    """📌 ดึงข้อมูล EC2 instances และแปลงเป็น JSON"""
    response = ec2_client.describe_instances()

    instances = []
    for reservation in response["Reservations"]:
        for instance in reservation["Instances"]:
            instance_data = {
                "InstanceID": instance["InstanceId"],
                "State": instance["State"]["Name"],
                "InstanceType": instance["InstanceType"],
                "PublicIP": instance.get("PublicIpAddress", "N/A"),
                "Name": next((tag["Value"] for tag in instance.get("Tags", []) if tag["Key"] == "Name"), "N/A"),
                "Project": next((tag["Value"] for tag in instance.get("Tags", []) if tag["Key"] == "Project"), "N/A")
            }
            instances.append(instance_data)

    return instances


def save_to_excel(instances):
    """📌 บันทึกข้อมูล JSON ลงไฟล์ Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "EC2 Instances"

    columns = ["InstanceID", "State", "InstanceType", "PublicIP", "Name", "Project"]
    ws.append(columns)

    # 🔹 ใส่ข้อมูลลงตาราง
    for instance in instances:
        row_data = [instance.get(col, "N/A") for col in columns]
        ws.append(row_data)

    # 🔹 จัดรูปแบบของเซลล์
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # 🔹 ไฮไลท์สีแดงถ้าเป็น "N/A" หรือไม่มีข้อมูล
            if cell.value in ["N/A", None, ""]:
                cell.fill = red_fill

    # 🔹 ปรับขนาดคอลัมน์ให้อัตโนมัติ
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    # 🔹 บันทึกไฟล์
    wb.save(TMP_FILE_PATH)
    return TMP_FILE_PATH


def upload_to_s3(file_path):
    """📌 อัปโหลดไฟล์ไปที่ S3"""
    s3_client.upload_file(file_path, S3_BUCKET_NAME, S3_FILE_NAME)
    s3_url = f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{S3_FILE_NAME}"
    return s3_url


def lambda_handler(event, context):
    """📌 ฟังก์ชันหลักของ Lambda"""
    instances = get_ec2_instances()
    excel_path = save_to_excel(instances)
    s3_link = upload_to_s3(excel_path)

    return {
        "statusCode": 200,
        "body": json.dumps(instances, indent=4),
        "s3_link": s3_link
    }
